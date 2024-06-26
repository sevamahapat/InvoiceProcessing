'''
@Author: Yao Xie
'''
from celery import shared_task
from google.oauth2.service_account import Credentials
import numpy as np
import pandas as pd
from pathlib import Path
from django.conf import settings
from rest_framework.response import Response
import os
import json
from . import prompt_engineering_blackbox as prompt

import openpyxl

# @shared_task
def generate_invoice_info_gpt(task_id):
    """
    This function generates the invoice information using GPT API
    Generates the parsed invoice information from passing the .pdf files to GPT API 
    """
    error_message = None
    file_name = f"{task_id}.pdf"  # Save the file with the task ID
    # Construct the file path
    file_path = Path(settings.MEDIA_ROOT) / file_name

    # check if the file exists
    if not os.path.exists(file_path):
        error_message = 'File ' + file_path + ' not found'
    
    # Pass the content to GPT API and get result
    result = None
    result = prompt.call(file_path)

    if result is None:
        error_message = 'Error processing the file: ' + file_path

    # print(result)
    # Find the position of the opening curly brace "{"
    brace_index = result.find("{")
    if brace_index == -1:
        error_message = 'No JSON like info when processing the file: ' + file_path
    
    # Extract the JSON-like string starting from the opening curly brace
    json_str = result[brace_index:]

    # Parse the JSON-like string into a dictionary
    result_data = json.loads(json_str)

    # Create a DataFrame from the dictionary
    result_df = pd.DataFrame([result_data])
    # for all columns name, remove the space
    result_df.columns = result_df.columns.str.rstrip()
    # if any column is missing, add it to the DataFrame
    if 'Invoice number' not in result_df:
        result_df['Invoice number'] = ''
    if 'Supplier Name' not in result_df:
        result_df['Supplier Name'] = ''
    if 'Supplier Address Street 1' not in result_df:
        result_df['Supplier Address Street 1'] = ''
    if 'Supplier Address Street 2' not in result_df:
        result_df['Supplier Address Street 2'] = ''
    if 'Supplier City' not in result_df:
        result_df['Supplier City'] = ''
    if 'Supplier State' not in result_df:
        result_df['Supplier State'] = ''
    if 'Supplier zip' not in result_df:
        result_df['Supplier zip'] = ''
    if 'Supplier Country' not in result_df:
        result_df['Supplier Country'] = ''
    if 'Ship To Street 1' not in result_df:
        result_df['Ship To Street 1'] = ''
    if 'Ship To Street 2' not in result_df:
        result_df['Ship To Street 2'] = ''
    if 'Ship To City' not in result_df:
        result_df['Ship To City'] = ''
    if 'Ship To State' not in result_df:
        result_df['Ship To State'] = ''
    if 'Zip' not in result_df:
        result_df['Zip'] = ''
    if 'Ship To Country' not in result_df:
        result_df['Ship To Country'] = ''
    if 'Invoice currency' not in result_df:
        result_df['Invoice currency'] = ''
    if 'Invoice amount (Incl tax)' not in result_df:
        result_df['Invoice amount (Incl tax)'] = 0
    if 'Invoice tax amount' not in result_df:
        result_df['Invoice tax amount'] = 0
    if 'Purchase Order' not in result_df:
        result_df['Purchase Order'] = ''
    return result_df
    # store the result into the sqlite database
    # new_invoice = Invoice(
    #     invoice_number = result_data['Invoice number'],
    #     supplier_name = result_data['Supplier Name'],
    #     supplier_address_street1 = result_data['Supplier Address Street 1'],
    #     supplier_address_street2 = result_data['Supplier Address Street 2'],
    #     supplier_city = result_data['Supplier City'],
    #     supplier_state = result_data['Supplier State'],
    #     supplier_zip = result_data['Supplier zip'],
    #     supplier_country = result_data['Supplier Country'],
    #     ship_to_street1 = result_data['Ship To Street 1'],
    #     ship_to_street2 = result_data['Ship To Street 2'],
    #     ship_to_city = result_data['Ship To City'],
    #     ship_to_state = result_data['Ship To State'],
    #     ship_to_zip = result_data['Zip'],
    #     ship_to_country = result_data['Ship To Country'],
    #     invoice_currency = result_data['Invoice currency'],
    #     invoice_amount_incl_tax = result_data['Invoice amount (Incl tax)'],
    #     invoice_tax_amount = result_data['Invoice tax amount'],
    #     purchase_order = result_data['Purchase Order'],
    # )
    # new_invoice.save()

    # # check if the output file 'result.xlsx' exists
    # file_path = Path(settings.MEDIA_ROOT) / 'result.xlsx'
    # if not os.path.exists(file_path):
    #     # Create a new Excel file
    #     result_df.to_excel(file_path, index=False)
    # else:
    #     # Load the existing Excel file
    #     workbook = openpyxl.load_workbook(file_path)

    #     # Select the first worksheet
    #     worksheet = workbook.active

    #     # Determine the next available row after the last written row
    #     next_row = worksheet.max_row + 1

    #     if error_message is not None:
    #         # Append the error message to the next available row
    #         worksheet.cell(row=next_row, column=1, value=error_message)
    #         return
    #     else:
    #         # Append the data to the next available row
    #         for index, row in result_df.iterrows():
    #             for col_index, value in enumerate(row, start=1):
    #                 worksheet.cell(row=next_row + index, column=col_index, value=value)

    #     # Save the changes to the Excel file
    #     workbook.save(file_path)

            

    
